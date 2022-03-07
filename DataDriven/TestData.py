import openpyxl

def DataGen():
    bk=openpyxl.load_workbook("C:\\Users\\srika\\OneDrive\\Desktop\\TestData.xlsx")
    sheet=bk['Sheet1']
    print(sheet)
    rw=sheet.max_row
    cl=sheet.max_column
    li=[]
    li1=[]
    for i in range(1,rw+1):
        li1=[]
        un=sheet.cell(i,1)
        up=sheet.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,up.value)
        li.insert(i-1,li1)
    print(li)

    return li


